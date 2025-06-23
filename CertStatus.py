import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv
import os
import logging
import hashlib
import time

load_dotenv()
ScriptDir = Path(__file__).parent
LogDir = ScriptDir / "A3Logs"
LogDir.mkdir(exist_ok=True)
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
LogFile = LogDir / f"log_{timestamp}.txt"

logging.basicConfig(
    filename=LogFile,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
start_time = time.time()
logging.info("Script started.")

def calculate_checksum(file_path):
    sha256 = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)
    return sha256.hexdigest()

EXLFile = Path(os.getenv('EXLFile'))
xl_checksum = calculate_checksum(EXLFile)
logging.info(f"Sumcheck SHA256: {xl_checksum}")

df = pd.read_excel(EXLFile, sheet_name="A3 Certificates")
df['Issue Date'] = pd.to_datetime(df['Issue Date'])
df['Due Date'] = pd.to_datetime(df['Due Date'])
today = datetime.today()
df = df.sort_values(by='Issue Date', ascending=False)
df_latest = df.drop_duplicates(subset=['Machinery', 'System'], keep='first').copy()
df_latest['Days Until Expiration'] = (df_latest['Due Date'] - today).dt.days

# Expiration status codes: 0=Not Expired, 1=Near, 2=Almost, 3=Expired, 4=Expired over a week
def expiration_code(days):
    if days < -7:
        return 4
    elif days < 0:
        return 3
    elif days < 7:
        return 2
    elif days < 15:
        return 1
    else:
        return 0

df_latest['Updated Notification'] = df_latest['Days Until Expiration'].apply(expiration_code)
df_latest['Notification'] = df_latest['Notification'].fillna(0).astype(int)
status_changed = df_latest[df_latest['Notification'] != df_latest['Updated Notification']]
status_changed = status_changed.sort_values(by='Days Until Expiration')

if not status_changed.empty:
    email_address = os.getenv('email_address')
    app_pws = os.getenv('app_pws')

    email_body = """
    <html>
    <body>
        <p>Καλημέρα,<br>
        Τα παρακάτω πιστοποιητικά εφιστούν την προσοχή σας:</p>

        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-family: Arial, sans-serif;">
        <thead>
            <tr>
            <th style="max-width: 100px; word-wrap: break-word;">Υπολειπόμενες Μέρες</th>
            <th style="max-width: 150px; word-wrap: break-word;">Μηχάνημα</th>
            <th style="max-width: 150px; word-wrap: break-word;">Σύστημα</th>
            <th style="max-width: 120px; word-wrap: break-word;">Ημ. Λήξης</th>
            <th style="max-width: 200px; word-wrap: break-word;">Τύπος Πιστοποίησης</th>
            <th style="max-width: 150px; word-wrap: break-word;">Αρ. Πιστοποίητικού</th>
            </tr>
        </thead>
        <tbody>
    """

    for _, row in status_changed.iterrows():

        notification_level = row['Updated Notification']
        if notification_level in [3, 4]:
            row_style = 'background-color: #f8d7da;'  # light red
        elif notification_level == 2:
            row_style = 'background-color: #fff3cd;'  # light yellow
        else:
            row_style = ''

        email_body += f"""
            <tr style="{row_style}">
            <td style="max-width: 100px; word-wrap: break-word;">{row['Days Until Expiration']}</td>
            <td style="max-width: 150px; word-wrap: break-word;">{row['Machinery']}</td>
            <td style="max-width: 150px; word-wrap: break-word;">{row['System']}</td>
            <td style="max-width: 120px; word-wrap: break-word;">{row['Due Date'].date()}</td>
            <td style="max-width: 200px; word-wrap: break-word;">{row['Certification Type']}</td>
            <td style="max-width: 150px; word-wrap: break-word;">{row['Certification No.']}</td>
            </tr>
        """

    email_body += """
        </tbody>
        </table>
    </body>
    </html>
    """
    msg = EmailMessage()
    msg['Subject'] = 'Ενημέρωση Πιστοποιητικών Μηχανημάτων'
    msg['From'] = email_address
    A3emails = os.getenv('A3emails')
    msg['To'] = ', '.join([email.strip() for email in A3emails.split(',')])
    msg.set_content("Το email απαιτεί πρόγραμμα ανάγνωσης HTML.")
    msg.add_alternative(email_body, subtype='html')

    try:

        with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            smtp.starttls()
            smtp.login(email_address, app_pws)
            smtp.send_message(msg)

        NotifCerts = ", ".join(f"{row['Machinery']} / {row['System']}" for _, row in status_changed.iterrows())
        logging.info(f"Email sent to: {A3emails} regarding: {NotifCerts}")

        wb = load_workbook(EXLFile)
        ws = wb['A3 Certificates']
        for _, row in status_changed.iterrows():
            excel_row = int(row['ID'])
            new_notification = int(row['Updated Notification'])
            ws.cell(row=excel_row, column=2).value = new_notification
        wb.save(EXLFile) 

        print("Email sent with updated certificate statuses.")
        
    except Exception as e:
        logging.error(f"Failed to send email or update file: {str(e)}")
else:
    logging.info("No changes detected. No email sent.")

end_time = time.time()
elapsed = round(end_time - start_time, 2)
logging.info(f"Script completed in {elapsed} seconds.")